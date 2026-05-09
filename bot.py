import json, time, re, schedule, logging, os, requests
from datetime import datetime
from io import BytesIO
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
from google.oauth2.service_account import Credentials
from google.auth.transport.requests import Request
from bs4 import BeautifulSoup

# ================= KONFİQURASİYA =================
EXCEL_FILE_URL = os.environ.get("EXCEL_FILE_URL", "")
PROSPECT_EXCEL_URL = os.environ.get("PROSPECT_EXCEL_URL", "") # YENİ: Anbar faylının linki
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")
PRICE_UNDERCUT = 0.01
MAX_WORKERS = 3 
# =================================================

# Umico Excel Sütunları (1-dən başlayaraq): F=6 (Say), H=8 (Qiymət), N=14 (URL), O=15 (Min), Q=17 (İD)
COL_SAY = 6; COL_QIYMET = 8; COL_URL = 14; COL_MIN = 15; COL_ID = 17

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "az,en-US;q=0.9,en;q=0.8"
})

def send_telegram(message):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID: return
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        payload = {"chat_id": TELEGRAM_CHAT_ID, "text": message, "parse_mode": "HTML"}
        session.post(url, json=payload, timeout=10)
    except Exception as e:
        log.error(f"Telegram xətası: {e}")

def send_telegram_document(file_bytes, filename, caption=""):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID: return
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
        data = {"chat_id": TELEGRAM_CHAT_ID, "caption": caption, "parse_mode": "HTML"}
        files = {"document": (filename, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        session.post(url, data=data, files=files, timeout=20)
    except Exception as e:
        log.error(f"Telegram sənəd xətası: {e}")

def parse_price(text):
    if not text: return 0.0
    cleaned = re.sub(r'[^0-9\.,]', '', str(text))
    if not cleaned: return 0.0
    if ',' in cleaned and '.' in cleaned: cleaned = cleaned.replace(',', '')
    elif ',' in cleaned: cleaned = cleaned.replace(',', '.')
    try: return round(float(cleaned), 2)
    except: return 0.0

def get_competitor_prices(url, product_name):
    competitors = []
    has_block = False
    try:
        resp = session.get(url, timeout=25)
        if resp.status_code != 200: return [], False
        
        html = resp.text
        if any(x in html.lower() for x in ["bütün satıcıların", "digər satıcılar", "bütün qiymətlər", "other-seller"]):
            has_block = True
            
        soup = BeautifulSoup(html, "html.parser")
        for tag in soup.find_all(attrs={"data-info": True}):
            if "price" in tag["data-info"].lower():
                p = parse_price(tag.get_text())
                if p > 0: competitors.append(p)
                
        seller_blocks = soup.find_all('div', class_=re.compile(r'other-seller', re.I))
        for block in seller_blocks:
             has_block = True
             price_tags = block.find_all(string=re.compile(r'[\d\.,\s]+(₼|AZN)', re.I))
             for tag in price_tags:
                 p = parse_price(tag)
                 if p > 0: competitors.append(p)
                 
        raw_prices = re.findall(r'["\']?price["\']?\s*[:=]\s*["\']?([\d\.,\s]+)["\']?', html, re.I)
        for p_str in raw_prices:
            p = parse_price(p_str)
            if p > 0: competitors.append(p)

    except Exception as e:
        log.error(f"Scraping xətası: {e}")
        
    return list(set(competitors)), has_block

def process_product(p):
    try:
        current = round(p['current'], 2)
        min_p = round(p['min'], 2)
        max_p = round(min_p * 1.05, 2) 
        
        all_found, has_block = get_competitor_prices(p['url'], p['name'])
        
        if not has_block:
            competitors = []
        else:
            # Taksit qoruyucusu
            competitors = [
                round(price, 2) for price in all_found 
                if price > (current * 0.6) and price < (max_p * 1.5) and abs(price - current) > 0.009
            ]
        
        # Hədəf qiyməti hesablamaq
        if not competitors:
            target = current # Rəqib yoxdursa qaldırma
        else:
            cheapest = min(competitors)
            target = max(cheapest - PRICE_UNDERCUT, min_p)
            target = min(target, max_p)

        # QİYMƏT QALDIRMAĞI QADAĞAN EDİRİK
        if target > current:
            target = current

        # Qiymət dəyişikliyini yoxla
        if abs(current - target) >= 0.009:
            emoji = "📉" if target < current else "📈"
            status_text = "Endirildi" if target < current else "Qaldırıldı"
            
            return {
                "status": "updated", 
                "row": p['row'], 
                "new": round(target, 2), 
                "name": p['name'], 
                "msg": f"{emoji} <b>{p['name']}</b>\nKöhnə: {current}₼ | Yeni: <b>{round(target, 2)}₼</b> ({status_text})"
            }
        
        if competitors and min(competitors) < current:
             return {
                "status": "limit_reached",
                "name": p['name'],
                "url": p['url'],
                "current": current,
                "competitor": min(competitors),
                "min": min_p,
                "max": max_p 
            }

        return {"status": "no_change", "name": p['name']}
            
    except Exception as e:
        return {"status": "error", "name": p['name'], "error": str(e)}

def run_check():
    log.info("🚀 Yoxlama başladı...")
    stats = {"total": 0, "updated": 0, "stock_updates": 0, "limit": 0, "error": 0, "no_change": 0}
    limit_reached_list = [] 
    updated_messages = []
    
    try:
        # Google Auth
        creds = Credentials.from_service_account_info(json.loads(os.environ.get("GOOGLE_CREDENTIALS", "{}")), 
                                                      scopes=["https://www.googleapis.com/auth/drive"])
        
        # 1. UMİCO EXCEL FAYLINI YÜKLƏYİRİK
        file_id = EXCEL_FILE_URL.split("/d/")[1].split("/")[0]
        resp = requests.get(f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx", timeout=30)
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        
        # ================= STOK (SAY) SİNXRONİZASİYASI =================
        if PROSPECT_EXCEL_URL:
            try:
                log.info("📦 Prospect Anbar məlumatları yoxlanılır...")
                p_file_id = PROSPECT_EXCEL_URL.split("/d/")[1].split("/")[0]
                p_resp = requests.get(f"https://docs.google.com/spreadsheets/d/{p_file_id}/export?format=xlsx", timeout=30)
                p_wb = openpyxl.load_workbook(BytesIO(p_resp.content), data_only=True)
                p_ws = p_wb.active
                
                # Anbar siyahısını yaradırıq {ID: SAY}
                stock_map = {}
                header_row_idx = 1
                for r_idx, row in enumerate(p_ws.iter_rows(values_only=True), 1):
                    if row and "ID nömrə" in str(row):
                        header_row_idx = r_idx
                        break
                
                for row in p_ws.iter_rows(min_row=header_row_idx+1, values_only=True):
                    if len(row) >= 4:
                        prod_id = str(row[2]).strip() if row[2] else None
                        qty_val = str(row[3]).strip() if row[3] is not None else "-"
                        
                        if prod_id:
                            # Tireni (-) 0 kimi qəbul edirik, qalanlarını rəqəmə çeviririk
                            parsed_qty = 0 if qty_val == '-' else int(float(qty_val))
                            stock_map[prod_id] = parsed_qty

                # Umico faylına stokları yazırıq
                for row_idx in range(2, ws.max_row + 1):
                    cell_id = ws.cell(row=row_idx, column=COL_ID).value
                    if cell_id:
                        prod_id = str(cell_id).strip()
                        if prod_id in stock_map:
                            current_qty = str(ws.cell(row=row_idx, column=COL_SAY).value).strip()
                            new_qty = stock_map[prod_id]
                            # Əgər rəqəm dəyişibsə yenilə
                            if current_qty != str(new_qty):
                                ws.cell(row=row_idx, column=COL_SAY, value=new_qty)
                                stats["stock_updates"] += 1
                                
                log.info(f"✅ Anbardan {stats['stock_updates']} məhsulun sayı Umico faylına yazıldı.")
            except Exception as e:
                log.error(f"❌ Stok sinxronizasiyası xətası: {e}")
        # ===============================================================

        # 2. QİYMƏT YOXLANMASI ÜÇÜN MƏHSULLARI SEÇİRİK
        products = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            url = row[COL_URL-1]
            if not url or "http" not in str(url): continue
            try:
                def f_val(v): return float(str(v or 0).replace(",",".").replace(" ","").replace("\xa0",""))
                products.append({"row": i, "url": str(url).strip(), "name": f"{row[3]} {row[2]}", "current": f_val(row[COL_QIYMET-1]), "min": f_val(row[COL_MIN-1])})
            except: continue

        stats["total"] = len(products)
        changes = []
        
        # Qiymətləri sürətlə yoxlayırıq
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = [executor.submit(process_product, p) for p in products]
            for f in as_completed(futures):
                res = f.result()
                if res["status"] == "updated":
                    changes.append(res)
                    stats["updated"] += 1
                    updated_messages.append(res["msg"])
                elif res["status"] == "limit_reached":
                    limit_reached_list.append(res)
                    stats["limit"] += 1
                elif res["status"] == "no_change":
                    stats["no_change"] += 1
                elif res["status"] == "error":
                    stats["error"] += 1

        # 3. YEKUN NƏTİCƏLƏRİ FAYLA YAZIB GOOGLE DRIVE-A YÜKLƏYİRİK
        if changes or stats["stock_updates"] > 0:
            for c in changes:
                ws.cell(row=c['row'], column=COL_QIYMET, value=c['new'])
            
            out = BytesIO()
            wb.save(out)
            creds.refresh(Request())
            requests.patch(f"https://www.googleapis.com/upload/drive/v3/files/{file_id}?uploadType=media",
                headers={"Authorization": f"Bearer {creds.token}"}, data=out.getvalue(), timeout=60)
            log.info(f"✅ Fayl Drive-da yeniləndi. (Qiymət: {len(changes)}, Stok: {stats['stock_updates']})")

        # Telegrama Qiymət Dəyişikliyi mesajları
        if updated_messages:
            chunk = "🔄 <b>Qiymət Güncəlləmələri:</b>\n\n"
            for msg in updated_messages:
                if len(chunk) + len(msg) > 3800:
                    send_telegram(chunk)
                    chunk = ""
                chunk += f"{msg}\n"
            if chunk.strip(): send_telegram(chunk)

        # Limit Exceli
        if limit_reached_list:
            wb_limit = openpyxl.Workbook()
            ws_limit = wb_limit.active
            ws_limit.title = "Limitə Çatanlar"
            ws_limit.append(["Məhsul Adı", "Bizim Qiymət", "Minimum Limit", "Maksimum Limit", "Ən Ucuz Rəqib", "Məhsul Linki"])
            for item in limit_reached_list:
                ws_limit.append([item["name"], item["current"], item["min"], item["max"], item["competitor"], item["url"]])
            
            out_limit = BytesIO()
            wb_limit.save(out_limit)
            out_limit.seek(0) 
            send_telegram_document(out_limit.read(), f"Limite_Dirananlar_{datetime.now().strftime('%d_%m_%H_%M')}.xlsx", "⚠️ Rəqib bizdən ucuzdur, amma limitə görə qiymət dəyişmədi.")

        # Yekun Hesabat
        report = (
            f"📊 <b>Yoxlama Hesabatı</b>\n"
            f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📦 Ümumi məhsul: <b>{stats['total']}</b>\n"
            f"🔢 Sayı (Stok) yeniləndi: <b>{stats['stock_updates']}</b>\n"
            f"🔄 Qiyməti endirildi: <b>{stats['updated']}</b>\n"
            f"⚠️ Limitə dirənən: <b>{stats['limit']}</b>\n"
            f"➖ Qiymət dəyişmədi: <b>{stats['no_change']}</b>\n"
            f"❌ Xəta: <b>{stats['error']}</b>"
        )
        send_telegram(report)

    except Exception as e:
        log.error(f"Sistem xətası: {e}")
        send_telegram(f"❌ <b>Sistem Xətası:</b>\n{str(e)}")

if __name__ == "__main__":
    run_check()
    schedule.every(10).minutes.do(run_check)
    while True:
        schedule.run_pending()
        time.sleep(1)
